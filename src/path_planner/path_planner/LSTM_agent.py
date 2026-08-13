from __future__ import annotations
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from collections.abc import Iterator, Sequence, Callable
    from gymnasium import Env

import shutil
import re
import torch
import torch.nn.functional as F
import torch.optim as optim
from torch.optim.lr_scheduler import CosineAnnealingLR
import wandb
import numpy as np
import matplotlib.pyplot as plt
import cv2
import time
import os
from dataclasses import asdict
from itertools import cycle
from gymnasium.vector import SyncVectorEnv
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

import glob
from IPython.display import display

# 앞서 정의한 클래스들을 임포트한다고 가정 (또는 같은 파일에 위치)
from path_planner.config import EnvConfig, TrainConfig, LOCAL_VIEW_DIM, RESULT_SAVE_DIR, TRAIN_MAP_FILE_REGEX, TEST_MAP_FILE_REGEX
from path_planner.map_generator import generate_multiple_maps
from path_planner.map_layer import MapConfigSchema
from path_planner.environment import CoverageEnv, ACTION_NUM
from path_planner.LSTM_network import LSTMPolicyNetwork
from path_planner.utils.utils import *
from path_planner.utils.visualizer import display_image
from path_planner.utils.trajectory_metrics import GRID_RESOLUTION_M, LINEAR_VELOCITY

class CustomGymVecEnv(SyncVectorEnv):
    """
    Environment 여러 개를 한번에 다룰 때 사용하는 class
    SyncVectorEnv 버전
    """
    
    def __init__(self, env_fns: Iterator[Callable[[], Env]] | Sequence[Callable[[], Env]], **kwargs):
        super().__init__(env_fns, **kwargs)

    def call_at(self, method_name: str="reset", env_id: int=0, *args, **kwargs):
        """
        특정 environment에만 개별적으로 method를 수행

        Args:
            method_name (str, optional): 개별 environment에서 불러온 method 이름. Defaults to "reset".
            env_id (int, optional): Method를 사용할 environment의 index. Defaults to 0.

        Returns:
            _type_: _description_
        """
        target_env = self.envs[env_id]
        attr = getattr(target_env, method_name)
        if callable(attr):
            return attr(*args, **kwargs)
        return attr
    
    def step(self, actions):
        """
        Step 함수 실행
        SyncVectorEnv에 정의된 step method는 환경 일부가 terminated or truncated 되었을 때 환경 reset을 수행할 것을 강요함.
        reset을 외부에서 직접 해주기 위해 이러한 기능을 제거한 step 함수를 새로 작성.
        """
        self._autoreset_envs.fill(False)
        return super().step(actions)


def is_better_path(cand: tuple[float, float, float], ref: tuple[float, float, float]) -> bool:
    """
    Path metric을 비교하는 함수
    1. Coverage가 90% 미만인 경우: Coverage가 높을수록 좋은 path
    2. Coverage가 90% 이상인 경우:
        1) Overlap 차이가 10%p 이상인 경우: Overlap 수치가 더 적은 path가 더 좋은 path
        2) Overlap 차이가 10%p 미만인 경우: 청소 시간이 더 적게 든 path가 더 좋은 path

    Args:
        cand (tuple[float, float, float]): 비교할 path의 (cov, overlap, time)
        ref (tuple[float, float, float]): 기준 path의 (cov, overlap, time)

    Returns:
        bool: cand의 path가 ref의 path보다 우수한 경우 True 반환
    """
    cov_c, ov_c, t_c = cand
    cov_r, ov_r, t_r = ref

    # Candidate과 reference의 coverage 90% 미만이면 coverage 우선으로 경로 비교
    if cov_c < 0.9 or cov_r < 0.9: 
        return cov_c > cov_r

    # 둘 다 coverage 90% 이상인 경우, overlap rate으로 경로 비교. 차이가 cleaning time으로 비교
    else:
        if abs(ov_c - ov_r) < 10.0: # 10%p 차이 미만이면 청소시간 비교
            return t_c < t_r
        return ov_c < ov_r

class LSTMAgent:
    
    def __init__(self, args):
        
        self.args = args
        self.seed = args.seed
        self.model_name = os.path.join(args.model_dir, args.model_name)
        self.map_save_dir = args.map_save_dir
        
        # Device 설정
        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        get_device_info(self.device) # Device 정보 출력
        
        # torch의 seed 설정
        set_torch_seed(args.seed)
        
        # Environment 조성
        env_args = {
            k: v for k, v in vars(args).items() 
            if k in EnvConfig.__dataclass_fields__ and v is not None
        }
        self.env_cfg = EnvConfig(**env_args)
        # Validation 및 test를 위한 environment 조성
        if args.mode == 'train':
            self.test_env_rng = np.random.default_rng(seed=args.seed+100000)
            self.test_map_rng = np.random.default_rng(seed=args.seed+100000)
            self.test_env = CoverageEnv(cfg=self.env_cfg, env_rng=self.test_env_rng, map_rng=self.test_map_rng) # 훈련용 environment와 다른 환경을 구성하기 위해서. validation용
        else:
            self.test_env_rng = np.random.default_rng(seed=args.seed+200000)
            self.test_map_rng = np.random.default_rng(seed=args.seed+200000)
            self.test_env = CoverageEnv(cfg=self.env_cfg, env_rng=self.test_env_rng, map_rng=self.test_map_rng)


        # Train map 및 validation map 생성
        if args.mode == 'train':
            self.train_map_sizes = ["40x40", "80x80", "Cropped"]
            if args.use_train_maps: # Train maps 생성
                self._check_and_generate_train_valid_maps(mode="train")
            self._check_and_generate_train_valid_maps(mode="valid") # Validation maps 생성

        # Test maps이 존재하는지 검토 후 저장
        elif args.mode == 'test':

            self.test_maps: dict[int, list[str]] = {}

            print(f"Checking if test maps are exist...")

            # Test map directory 검사
            test_maps_dir = os.path.join(args.map_save_dir, args.test_map_folder_name)
            if not os.path.isdir(test_maps_dir):
                raise FileNotFoundError(f"Test map directory '{test_maps_dir}' was not found. Please make sure the test maps are prepared before evaluation.")

            # Test map files 검사
            test_map_npy_files_name = sorted([
                f for f in os.listdir(test_maps_dir) 
                if f.endswith('.npy') and os.path.isfile(os.path.join(test_maps_dir, f))
            ])
            if len(test_map_npy_files_name) > 0:
                print(f"  There are {len(test_map_npy_files_name)} map files for testing.")
            else:
                raise FileNotFoundError(f"There is no test map is folder: {test_maps_dir}")

            for file_name in test_map_npy_files_name:
                full_path = os.path.join(test_maps_dir, file_name)
                match = re.search(TEST_MAP_FILE_REGEX, file_name)
                gd = match.groupdict()
                level = gd['level']
                if level not in self.test_maps.keys():
                    self.test_maps[level] = []
                self.test_maps[level].append(full_path)
        
            # Map 순서 정렬
            for key in self.test_maps.keys():
                self.test_maps[key].sort()


        # Policy network 조성
        ang_diff_diridx = self.test_env.ang_diff_diridx
        if ang_diff_diridx is None:
            raise ValueError("PolicyNetwork를 초기화하려면 'ang_diff_diridx' 인자가 반드시 필요합니다")
        network_config = {
            'action_enc_dim': args.action_enc_dim,
            'locel_view_dim': LOCAL_VIEW_DIM,
            'stack_steps': self.env_cfg.stack_steps,
            'map_feat_dim': 512,
            'loc_vec_in_dim': ACTION_NUM*2,
            'vec_feat_dim': 64,
            'lstm_in_dim': 5,
            'lstm_in_prj_dim': args.lstm_in_prj_dim,
            'lstm_hid_dim': args.lstm_hid_dim,
            'ang_diff_diridx': ang_diff_diridx,
        }
        self.policy_net = LSTMPolicyNetwork(**network_config).to(self.device)

        
        # Train과 관련된 instance 변수는 mode가 train일 때만 생성
        if args.mode == 'train':
            
            # Train hyperparameter를 받음
            train_args = {
                k: v for k, v in vars(args).items() 
                if k in TrainConfig.__dataclass_fields__ and v is not None
            }
            self.train_cfg = TrainConfig(**train_args)
            
            # Training용 environment 목록: Batch size 만큼의 environment를 조성하고 training data를 한번에 생성
            num_envs = self.train_cfg.batch_size
            self.train_env_rngs: list[np.random.Generator] = [] 
            self.train_map_rngs: list[np.random.Generator] = []
            for i in range(num_envs):
                self.train_env_rngs.append(np.random.default_rng(seed=args.seed+i))
                self.train_map_rngs.append(np.random.default_rng(seed=args.seed+i))
            env_fns = [lambda i=i: CoverageEnv(cfg=self.env_cfg, env_rng=self.train_env_rngs[i], map_rng=self.train_map_rngs[i]) for i in range(num_envs)]

            self.train_envs = CustomGymVecEnv(env_fns, autoreset_mode="Disabled")

            ########## Train map 설정 관련 iterator: Iterator, count, len이 세트 ##########
            # Train할 map_size를 결정하는 iterator
            self.train_maps_sizes_iter = iter(self.train_map_sizes)
            self.train_maps_sizes_iter_count = 0

            # Train할 map_size에서 어떤 level의 map을 뽑을지 결정하는 iterator
            self.train_maps_level_iter = {}
            self.train_maps_level_iter_len = {}
            self.train_maps_level_iter_counts = {}
            for map_size in self.train_map_sizes:
                self.train_maps_level_iter[map_size] = cycle(range(1, 5)) # 각 map size마다 level iterator를 생성
                self.train_maps_level_iter_len[map_size] = 4
                self.train_maps_level_iter_counts[map_size] = 0

            # args.use_train_maps가 참인 경우, self.train_maps에 iterator 저장
            self.train_maps_iter_len = {}
            self.train_maps_iter_counts = {}
            if args.use_train_maps:
                # Training에 필요한 map 얻기
                self.train_maps = self._get_map_dict(mode="train") # Key: Map size, Value: {Key=level: Value=[map file list]}
            ##############################################################################

            # else:
            #     self.train_map_sizes = []
            #     window_size = self.env_cfg.map_cfg.window_size
            #     for mult in range(1, 5):
            #         map_size = (window_size*mult, window_size*mult)
            #         self.train_map_sizes.append(map_size)
            #     self.train_map_sizes.append('None')

            # Validation에 필요한 map 얻기
            self.valid_maps = self._get_map_dict(mode="valid") # Key: Map size, Value: {Key=level: Value=[map file list]}
                    
            # Training에 필요한 변수 설정
            self.total_steps = 0            # Training을 진행한 steps 수 (Warmup 과정 포함)
            self.start_episode = 1
            self.eps_num_per_map_size = 0   # 주어진 map size에서 training한 episode 개수
            
            # Validation에 필요한 변수 설정
            self.max_coverage_mean = 0.0                # Validation을 수행했을 때 가장 높았던 coverage
            self.min_overlap_percent_mean = float('inf')   # Validation을 수행했을 때 가장 낮았던 overlap rate
            self.min_cleaning_time_mean = float('inf')  # Validation을 수행했을 때 가장 낮았던 cleaning time    
            self.best_traj_img = None                   # Validation을 수행했을 때 가장 coverage를 잘 했던 trajectory를 img로 저장 (RGB image)
            # Validation reward scales differ by curriculum map size, so best
            # checkpoints must be compared only within the same map size.
            self.best_validation_by_map_size = {
                map_size: {
                    'path_reward': float('-inf'),
                    'coverage': 0.0,
                    'overlap_percent': float('inf'),
                    'cleaning_time': float('inf'),
                }
                for map_size in self.train_map_sizes
            }
            # self._resume_curr_map_size = None
            # self._resume_no_more_map_size = False
            
            # Training을 위한 난수 생성기의 seed 설정
            self.train_rng = np.random.default_rng(seed=args.seed)
            
            # Training process를 볼 logger 설정: wandb
            self._setup_logging()
            
            # Optimizer 설정
            if self.train_cfg.optimizer == 'sgd':
                self.optimizer = optim.SGD(self.policy_net.parameters(), lr=self.train_cfg.lr, momentum=self.train_cfg.momentum)
            elif self.train_cfg.optimizer == 'adam':
                self.optimizer = optim.Adam(self.policy_net.parameters(), lr=self.train_cfg.lr)
            else:
                raise ValueError(f"Unsupported optimizer type: {self.train_cfg.optimizer}")

            self.scheduler = CosineAnnealingLR(self.optimizer, T_max=self.train_cfg.scheduler_max_step, eta_min=1e-6)

        # Loading model
        self._load_model(args)
         
    def _check_and_generate_train_valid_maps(self, mode: str="train"):

        if mode == "train":
            mode_str = "training"
        elif mode == "valid":
            mode_str = "validation"
        else:
            raise ValueError(f"Invalid mode {mode} at _check_and_generate_train_valid_maps method: mode must be \"train\" or \"valid\"")
        print(f"Checking pre-generated maps for {mode_str}...")
                    
        maps_folder = os.path.join(self.args.map_save_dir, mode)
        make_maps = False
        
        # 1. Map을 저장하는 폴더가 있는지 검사
        if not os.path.isdir(maps_folder):
            print(f"  Folder '{maps_folder}' does not exist.")
            make_maps = True
        else:
            # 2. 각 사이즈별 map이 모두 존재하는지 검사
            for map_size_str in self.train_map_sizes:
                sub_folder_path = os.path.join(maps_folder, map_size_str)
                if not os.path.isdir(sub_folder_path):
                    print(f"  Map size folder '{sub_folder_path}' does not exist.")
                    make_maps = True
                    break

                npy_files = glob.glob(os.path.join(sub_folder_path, "**", "*.npy"), recursive=True)
                map_count = len(npy_files)
                if map_count == 0:
                    print(f"  There is no map with map size {map_size_str}.")
                    make_maps = True
                    break
                else:
                    print(f"  There are {map_count} maps with map size {map_size_str}")

        # 3. 각 map size 폴더에 map이 하나라도 없을 경우, map 전체를 재생성
        if make_maps:
            print(f"Regenerating {mode_str} maps...")

            if os.path.exists(maps_folder):
                shutil.rmtree(maps_folder)

            for map_size_str in self.train_map_sizes:
                map_folder_name = os.path.join(maps_folder, map_size_str)
                os.makedirs(map_folder_name, exist_ok=True)

                if map_size_str == "Cropped":
                    map_size = None
                    start_mode = "corner"
                else:
                    H, W = map(int, map_size_str.split('x'))
                    map_size = (H, W)
                    start_mode = "edge"

                if mode == "train":
                    map_num = 1000 if map_size_str == "Cropped" else 200
                else:
                    map_num = self.args.valid_map_num
                print(f"  Generating {map_num} maps per level with map size {map_size_str}...")
                generate_multiple_maps(
                    map_folder_name=map_folder_name,
                    robot_diameter=self.env_cfg.robot_size, 
                    mode=mode, 
                    start_mode=start_mode,
                    seed=self.seed,
                    map_num_per_cond=map_num, 
                    visualize=True, 
                    map_size=map_size
                )
                print(f"  Complete!")
        else:
            print(f"Using pre-generated maps for {mode_str}.\n")


    def _get_map_dict(self, mode: str="train"):

        map_dict = {}

        map_dir = os.path.join(self.map_save_dir, mode)
                        
        # 각 map의 path를 level, map size별로 나누어서 저장
        for folder in self.train_map_sizes:
            folder_path = os.path.join(map_dir, folder)
            map_dict[folder] = {}
            if os.path.exists(folder_path):
                # Image 폴더는 제외하고 .npy 파일들만 정확하게 수집
                npy_file_names = sorted([
                    f for f in os.listdir(folder_path) 
                    if f.endswith('.npy') and os.path.isfile(os.path.join(folder_path, f))
                ])
                
                for file_name in npy_file_names:
                    full_path = os.path.join(folder_path, file_name)
                    match = re.search(TRAIN_MAP_FILE_REGEX, file_name)
                    gd = match.groupdict()
                    level = int(gd['level'])
                    if level not in map_dict[folder]:
                        map_dict[folder][level] = []
                    map_dict[folder][level].append(full_path)
                
                for level in map_dict[folder].keys():
                    map_dict[folder][level].sort()
                    if mode == "train":
                        self.train_maps_iter_len.setdefault(folder, {})[level] = len(map_dict[folder][level])
                        self.train_maps_iter_counts.setdefault(folder, {})[level] = 0
                        map_dict[folder][level] = cycle(map_dict[folder][level])

        return map_dict


    def _setup_logging(self):
        """
        Training process를 지켜 볼 tool 설정: wandb
        """
        import pytz
        import datetime
        
        # 현재 시간 얻기
        kst = pytz.timezone('Asia/Seoul')
        current_time = datetime.datetime.now(kst).strftime("%y%m%d_%H%M")
        
        # Training 조건을 구별하기 위해 표시할 hyperparmeter 설정
        params_list_for_log = [
            'batch_size', 'lr', 'min_lr', 'scheduler_max_step', 'optimizer', 'momentum', 
            'max_step_per_eps', 'max_eps_per_map_size',
            'action_enc_dim', 'lstm_in_prj_dim', 'lstm_hid_dim', 'detach_period'    
        ]
        train_cfg_dict = asdict(self.train_cfg)
        params_config = {k: v for k, v in train_cfg_dict.items() if k in params_list_for_log}
            
        # wandb 설정
        self.wandb_run = None
        if self.args.use_wandb:
            self.wandb_run = wandb.init(
                entity="lg-robot-cleaner",
                project="Robot_vacuum_ReDQN",
                config=params_config,
                name=f"{current_time}_{self.args.model_name}_training"
            )
    
    def _load_model(self, args):
        """
        Train:     
        - model_name: 최종적으로 훈련시킨 모델의 이름
        - checkpoint: 훈련 도중 저장되는 중간 결과물들은 (모델 이름)_checkpoints/ 폴더에 저장됨.
        - Checkpoint 파일명: (모델 이름)_(에피소드 번호).pth
        - 모델 저장 경로 설정: 모델은 models/ 디렉토리에 저장됨.
        - 하위 폴더는 checkpoint를 모으는 폴더로 training 과정에 저장됨.
        - 훈련이 다 된 모델은 models/ 폴더 바로 아래에 있음.
        
        Test:
        - model_name: Test할 모델 이름
        """
        model_dir = args.model_dir # Model이 담긴 폴더: .../models
        model_name_base, _ = os.path.splitext(args.model_name) # 확장자 '.pth'를 제거한 model_name
        checkpoint_dir = os.path.join(model_dir, model_name_base + "_checkpoints") # Checkpoint 저장 폴더
        
        # 폴더 생성
        os.makedirs(model_dir, exist_ok=True)
        os.makedirs(checkpoint_dir, exist_ok=True)
        
        if args.mode == 'train':
            
            model_path = None
            is_checkpoint = False
            
            # 1. Loading할 model 선택
            if args.pre_model_name: # 사전 학습 완료된 모델부터 이어서 훈련하는 경우
                print("Use pre-trained model...")
                model_path = os.path.join(model_dir, args.pre_model_name)
                print(f"Loading pre-trained model: {model_path}")
            else: # 사전 학습 완료된 모델이 없는 경우(처음부터 훈련하는 경우 + checkpoint에서 이어서 훈련하는 경우)
                checkpoints = glob.glob(os.path.join(checkpoint_dir, "*.pth"))
                if checkpoints:
                    print("Checkpoint directory already exists. Continue training...")
                    model_path = max(checkpoints, key=os.path.getctime) # 가장 최근 checkpoint 파일부터 훈련 재개
                    is_checkpoint = True
                    print(f"Loading latest checkpoint: {model_path}")
                else:
                    print("No checkpoint files found!")
                    return
                
            # 2. Model loading
            if not os.path.isfile(model_path):
                raise FileNotFoundError(f"Model file not found: {model_path}")
            data = torch.load(model_path, map_location=self.device, weights_only=False)
            self.policy_net.load_state_dict(data['model_state_dict'])
            
            if is_checkpoint: # Checkpoint를 사용하는 경우 optimizer와 step 수도 추가로 load
                for map_size, valid_metric_dict in data.get('best_validation_by_map_size', {}).items():
                    if map_size in self.best_validation_by_map_size:
                        self.best_validation_by_map_size[map_size].update(valid_metric_dict)

                # Optimizer loading
                try:
                    self.optimizer.load_state_dict(data['optimizer_state_dict']) # Optimizer도 checkpoint와 똑같이 유지
                except KeyError as e:
                    print(f"Optimizer state mismatch. Skipping: {e}")

                # Scheduler loading
                try:
                    self.scheduler.load_state_dict(data['scheduler_state_dict']) # Optimizer도 checkpoint와 똑같이 유지
                except KeyError as e:
                    print(f"Scheduler state mismatch. Skipping: {e}")
              
                # 추가적인 변수 loading
                self.total_steps = data['total_steps']

                self.train_maps_sizes_iter_count = data.get('train_maps_sizes_iter_count', 0)
                self.train_maps_level_iter_counts.update(data.get('train_maps_level_iter_counts', {}))
                if args.use_train_maps:
                    for map_size, counts in data.get('train_maps_iter_counts', {}).items():
                        if map_size in self.train_maps_iter_counts:
                            self.train_maps_iter_counts[map_size].update(counts)

                self.start_episode = data['episode'] + 1
                self.eps_num_per_map_size = data['eps_num_per_map_size']

                best_traj_img_path = os.path.join(checkpoint_dir, args.best_traj_img_name)
                if os.path.isfile(best_traj_img_path):
                    img = cv2.imread(best_traj_img_path) # BGR image
                    self.best_traj_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB) # BGR -> RGB 변환

                # seed 진행도 복원
                for rng, state in zip(self.train_env_rngs, data['train_env_rng_states']):
                    rng.bit_generator.state = state
                for rng, state in zip(self.train_map_rngs, data['train_map_rng_states']):
                    rng.bit_generator.state = state
                self.train_rng.bit_generator.state = data['train_rng_state']
                self.test_env_rng.bit_generator.state = data['test_env_rng_state']
                self.test_map_rng.bit_generator.state = data['test_map_rng_state']
 
                cpu_rng_state = data['torch_cpu_rng_state']
                if isinstance(cpu_rng_state, torch.Tensor):
                    cpu_rng_state = cpu_rng_state.to(dtype=torch.uint8, device='cpu')
                torch.set_rng_state(cpu_rng_state)

                if torch.cuda.is_available() and data.get('torch_gpu_rng_states') is not None:
                    gpu_rngs = [
                        s.to(dtype=torch.uint8, device='cpu') if isinstance(s, torch.Tensor) else s 
                        for s in data['torch_gpu_rng_states']
                    ]
                    torch.cuda.set_rng_state_all(gpu_rngs)

                # Map 설정 iterator 상태 복원
                # consumed_map_sizes = self.train_maps_sizes_iter_count + (1 if self._resume_curr_map_size is not None else 0)
                for _ in range(self.train_maps_sizes_iter_count):
                    try:
                        next(self.train_maps_sizes_iter)
                    except StopIteration:
                        break

                for map_size, count in self.train_maps_level_iter_counts.items():
                    for _ in range(count % self.train_maps_level_iter_len[map_size]):
                        next(self.train_maps_level_iter[map_size])

                if self.train_cfg.use_train_maps:
                    for map_size, count_dict in self.train_maps_iter_counts.items():
                        for level, count in count_dict.items():
                            map_len = self.train_maps_iter_len[map_size][level]
                            for _ in range(count % map_len):
                                next(self.train_maps[map_size][level])

                # Config 확인
                saved_env_cfg = data.get('env_config')
                saved_train_cfg = data.get('train_config')
                if saved_env_cfg is not None and saved_env_cfg != asdict(self.env_cfg):
                    print('Warning: EnvConfig differs from checkpoint; exact resume is not guaranteed.')
                if saved_train_cfg is not None and saved_train_cfg != asdict(self.train_cfg):
                    print('Warning: TrainConfig differs from checkpoint; exact resume is not guaranteed.')
                
                print(f"Resumed training from episode {data['episode']}")
                
        else:
            test_model = os.path.join(model_dir, args.model_name)
            if not os.path.exists(test_model):
                raise FileNotFoundError(f"Test model file not found: {test_model}")
            checkpoint = torch.load(test_model, map_location=self.device, weights_only=False)
            self.policy_net.load_state_dict(checkpoint['model_state_dict'])
            print(f"Test model loaded: {test_model}")
    
    def _save_model(self, mode='model', info=None):
        
        model_dir = self.args.model_dir # Model이 담긴 폴더명: .../models
        model_name_base, _ = os.path.splitext(self.args.model_name)
        checkpoint_dir = os.path.join(self.args.model_dir, model_name_base + "_checkpoints") # Checkpoint 저장 폴더
        
        os.makedirs(model_dir, exist_ok=True)
        os.makedirs(checkpoint_dir, exist_ok=True)
        
        save_path = None
        save_data = {
            'model_state_dict': self.policy_net.state_dict(),
            'best_validation_by_map_size': self.best_validation_by_map_size,
        }

        if mode == 'model': # model을 /models에 저장
            save_path = os.path.join(self.args.model_dir, self.args.model_name) # model 저장 경로
            
            # Checkpoint 파일에 가장 coverage가 잘 수행된 trajectory 그림을 저장
            if self.best_traj_img is not None:
                img_file_path = os.path.join(checkpoint_dir, self.args.best_traj_img_name)
                img_bgr = cv2.cvtColor(self.best_traj_img, cv2.COLOR_RGB2BGR) # OpenCV 저장용: (RGB -> BGR)
                cv2.imwrite(img_file_path, img_bgr)
                
        elif mode == 'checkpoint': # checkpoint 폴더에 저장
            
            assert info is not None
            episode = info['episode']
            save_path = os.path.join(checkpoint_dir, model_name_base + f"_{episode}.pth")
            save_data.update({
                'episode': episode,
                'total_steps': self.total_steps,
                'optimizer_state_dict': self.optimizer.state_dict(),
                'scheduler_state_dict': self.scheduler.state_dict(),
                'train_maps_sizes_iter_count': self.train_maps_sizes_iter_count,
                'train_maps_level_iter_counts': self.train_maps_level_iter_counts,
                'train_maps_iter_counts': self.train_maps_iter_counts,
                'eps_num_per_map_size': self.eps_num_per_map_size,
                'env_config': asdict(self.env_cfg),
                'train_config': asdict(self.train_cfg),

                # seed 진행도 저장
                'train_env_rng_states': [rng.bit_generator.state for rng in self.train_env_rngs],
                'train_map_rng_states': [rng.bit_generator.state for rng in self.train_map_rngs],
                'train_rng_state': self.train_rng.bit_generator.state,
                'test_env_rng_state': self.test_env_rng.bit_generator.state,
                'test_map_rng_state': self.test_map_rng.bit_generator.state,
                'torch_cpu_rng_state': torch.get_rng_state(),
                'torch_gpu_rng_states': torch.cuda.get_rng_state_all() if torch.cuda.is_available() else None,
            })
            
        else:
            raise ValueError(f"Unsupported mode: {mode}. Expected 'model' or 'checkpoint'.")

        torch.save(save_data, save_path)    
        
    def _pre_process_obs(self, obs_list, local_view_dim=LOCAL_VIEW_DIM) -> dict:
        
        """
        여러 환경의 obs들이 담긴 obs_list를 받아 병렬적으로 전처리하여 tensor로 변환.
        local_view_map은 적절한 크기로 resize만 진행.
        """
        # 1. CPU 리스트 데이터를 빠르게 PyTorch GPU 텐서로 통합 (연산 그래프 분리 .detach())
        if isinstance(obs_list, dict):
            maps = torch.as_tensor(obs_list['map'], device=self.device).float().detach() # Shape: (B, C, H, W) or (C, H, W)
            loc_vecs = torch.as_tensor(obs_list['loc_vec'], device=self.device).float().detach() # Shape: (B, vector_dim) or (vector_dim,)
            glob_vecs = torch.as_tensor(obs_list['glob_vec'], device=self.device).float().detach() # Shape: (B, vector_dim) or (vector_dim,)
            action_masks = torch.as_tensor(obs_list['action_mask'], device=self.device).float().detach() # Shape: (B, action_num) or (action_num,)

            if maps.ndim == 3: # local map view가 (C, H, W)로 들어오는 경우
                maps = maps.unsqueeze(0)
                loc_vecs = loc_vecs.unsqueeze(0)
                glob_vecs = glob_vecs.unsqueeze(0)
                action_masks = action_masks.unsqueeze(0)

        elif isinstance(obs_list, (list, tuple)):
            maps = (torch.stack([torch.as_tensor(o["map"], device=self.device) for o in obs_list], dim=0).float().detach()) # Shape: (B, C, H, W)
            loc_vecs = (torch.stack([torch.as_tensor(o["loc_vec"], device=self.device) for o in obs_list], dim=0).float().detach()) # Shape: (B, vector_dim)
            glob_vecs = (torch.stack([torch.as_tensor(o["glob_vec"], device=self.device) for o in obs_list], dim=0).float().detach()) # Shape: (B, vector_dim)
            action_masks = (torch.stack([torch.as_tensor(o["action_mask"], device=self.device) for o in obs_list], dim=0).float().detach()) # Shape: (B, action_num)
        else:
            raise TypeError(f"Unexpected type for obs_list: {type(obs_list)}. Expected dict or list.")
        
        # 2. local_view_map resize
        B, C, H, W = maps.shape
        target_dim = local_view_dim
        if H != target_dim or W != target_dim:
            total_patch = F.interpolate(maps, size=(target_dim, target_dim), mode='area')
        else:
            total_patch = maps

        return {
            "map": total_patch,           # Shape: (B, C, local_view_dim, local_view_dim)
            "loc_vec": loc_vecs,          # Shape: (B, loc_vec_dim)
            "glob_vec": glob_vecs,        # Shape: (B, glob_vec_dim)
            "action_mask": action_masks   # Shape: (B, action_num)
        }

    @staticmethod
    def _calculate_path_reward(
        coverage: torch.Tensor,
        overlap_rate: torch.Tensor,
        cleaning_time_rel: torch.Tensor,
        map_size: str,
    ) -> torch.Tensor:
        """Return the curriculum path reward used by both training and validation.

        ``overlap_rate`` must be a fraction in [0, 1], not a percentage.
        """
        cov_rwd_max = 0.5
        
        if map_size == '40x40':
            cov_thres = 0.9
            above_cov_thres_mask = (coverage >= cov_thres).float()
            cov_reward = torch.clamp(coverage, min=0.0, max=cov_thres) * (cov_rwd_max / cov_thres)

            overlap_scale = torch.clamp(1.0 - overlap_rate**2, min=-0.1)
            overlap_reward = overlap_scale * (1.0 - cov_rwd_max)

            time_penalty = torch.clamp((cleaning_time_rel - 1.0) * 0.025, min=0.0, max=0.1)

            return cov_reward + (overlap_reward - time_penalty) * above_cov_thres_mask

        cov_thres = 0.75
        above_cov_thres_mask = (coverage >= cov_thres).float()
        cov_reward = torch.clamp(coverage, min=0.0, max=cov_thres) * (cov_rwd_max / cov_thres)

        r_quad = (200.0 / 9.0) * torch.square(torch.clamp(coverage - cov_thres, min=0.0))
        overlap_max_weight = torch.clamp(r_quad, min=0.0, max=1.0 - cov_rwd_max)
        overlap_scale = torch.clamp(1.0 - torch.square(overlap_rate), min=-1.0)
        overlap_reward = overlap_scale * overlap_max_weight

        time_penalty = torch.clamp((cleaning_time_rel - 1.0) * 0.25, min=0.0, max=0.1)
        scaled_time_penalty = time_penalty * (overlap_max_weight / (1.0 - cov_rwd_max))
        return cov_reward + (overlap_reward - scaled_time_penalty) * above_cov_thres_mask
           
    def _validation(self, episode: int, curr_map_size: str, start_mode: str = "corner"):
        
        self.policy_net.eval() # eval mode로 전환
        coverage = []; overlap_percent = []; cleaning_time = []
        path_reward = []
        best_single_path_reward = float('-inf')
        coverage_mean = 0.0; overlap_percent_mean = 0.0; cleaning_time_mean = 0.0
        best_traj_img = None # 가장 성능이 좋았던 map에서의 trajectory를 보여줌
        
        print(f"Validation on {curr_map_size} map size")

        # Key: Map size, Value: {Key=level: Value=[map file list]}
        # Model의 성능 평가
        for level in range(1, 5):

            for i in range(self.train_cfg.valid_map_num):

                map_path = self.valid_maps[curr_map_size][level][i]
                map_config = MapConfigSchema(file_path=map_path)
                
                # Validation 환경을 초기화
                reset_info = self.test_env.reset(start_mode=start_mode, map_config=map_config)
                if reset_info:
                    obs, _ = reset_info
                else:
                    break # 다음 map_config를 이용
                    
                for start_num in range(self.train_cfg.valid_start_point_num):
                    
                    # start_num == 0인 경우 map을 초기화하면서 시작 지점도 초기화가 되었으므로 reset을 실행하지 않음.
                    if start_num != 0:
                        obs, _ = self.test_env.reset(start_mode=start_mode) # 시작 지점 초기화
                    cur_coverage, cur_overlap_percent, cur_cleaning_time, _ = self._test_one_map(self.test_env, obs, debug=False) # Coverage 성능을 평가

                    total_steps = self.test_env.steps
                    straight_cleaning_time = total_steps * GRID_RESOLUTION_M / LINEAR_VELOCITY / 60.0
                    cleaning_time_rel = cur_cleaning_time / (straight_cleaning_time + 1e-8)
                    cur_path_reward = self._calculate_path_reward(
                        torch.tensor(cur_coverage, device=self.device),
                        torch.tensor(cur_overlap_percent / 100.0, device=self.device),
                        torch.tensor(cleaning_time_rel, device=self.device),
                        curr_map_size,
                    ).item()

                    coverage.append(cur_coverage)
                    overlap_percent.append(cur_overlap_percent)
                    cleaning_time.append(cur_cleaning_time)
                    path_reward.append(cur_path_reward)

                    if cur_path_reward > best_single_path_reward:
                        best_single_path_reward = cur_path_reward
                        best_traj_img = self.test_env.get_visualized_img()
                    
        coverage_mean = np.mean(coverage) if coverage else 0.0
        overlap_percent_mean = np.mean(overlap_percent) if overlap_percent else 0.0
        cleaning_time_mean = np.mean(cleaning_time) if cleaning_time else 0.0
        path_reward_mean = np.mean(path_reward) if path_reward else 0.0
        
        curr_metrics = {
            'path_reward': path_reward_mean,
            'coverage': coverage_mean,
            'overlap_percent': overlap_percent_mean,
            'cleaning_time': cleaning_time_mean,
        }
        best_metrics = self.best_validation_by_map_size.setdefault(
            curr_map_size,
            {'path_reward': float('-inf'), 'coverage': 0.0,
             'overlap_percent': float('inf'), 'cleaning_time': float('inf')},
        )
        saved_best_model = curr_metrics['path_reward'] > best_metrics['path_reward']
        if saved_best_model:
            self.best_validation_by_map_size[curr_map_size].update(curr_metrics)
            if best_traj_img is not None:
                self.best_traj_img = best_traj_img
            self._save_model(mode='model')
            print(f"[Model Saved] New best path reward achieved. Model saved! / Past best reward: {best_metrics['path_reward']}, Current best reward: {curr_metrics['path_reward']}")
        
        # Validation 결과를 wandb에 기록
        if self.wandb_run:
            self.wandb_run.log({'Validation/Coverage_mean': coverage_mean,
                                'Validation/Overlap Rate Mean': overlap_percent_mean,
                                'Validation/Cleaning Time Mean': cleaning_time_mean,
                                'Validation/Path Reward Mean': path_reward_mean,
                                'Validation/Best Model Saved': int(saved_best_model),
                                'Validation/Best_path': wandb.Image(best_traj_img)}, step=self.total_steps)
        
        # Validation 결과 출력
        print(f"[Validation] Episode {episode}: \n"
              f"\tCoverage Mean = {coverage_mean*100:.2f}%\n"
              f"\tOverlap Rate Mean = {overlap_percent_mean:.2f}%\n"
              f"\tCleaning Time Mean = {cleaning_time_mean:.2f} min\n"
              f"\tPath Reward Mean = {path_reward_mean:.2f}")
            
        self.policy_net.train() # train mode로 전환

        return path_reward_mean
    
    def _test_one_map(self, env: CoverageEnv, obs: dict, 
                      target_cov_list: list[float] = [0.85, 0.90, 0.95],
                      debug: bool = False) -> tuple[float, float, float, dict[float, tuple[float, float]]]:

        # 1. Target coverage를 달성할 때마다 청소 시간을 저장. 달성하지 못한 경우는 -1로 저장하여 구별.
        sorted_target_cov = sorted(target_cov_list)
        target_cov_time_dict = {cov: (-1.0, -1.0) for cov in target_cov_list}
        target_idx = 0
        
        # 2. Path 생성: LSTM network에 통과시켜서 각 action의 확률값을 얻고 map 상의 전체 경로 생성
        done = False
        debug_skip_count = 0
        
        h_t = torch.zeros(1, self.policy_net.lstm_hid_dim, device=self.device)
        c_t = torch.zeros(1, self.policy_net.lstm_hid_dim, device=self.device)
        
        processed_obs = self._pre_process_obs(obs, local_view_dim=LOCAL_VIEW_DIM)

        # LSTM의 hidden state와 cell state, processed_obs를 복원하기 위한 list
        state_history = []
        
        # 디버그 모드일 때 사용할 도화지(fig)를 미리 딱 한 번만 만듦
        if debug:
            fig, axes = plt.subplots(2, 1, figsize=(15, 20))
            # 초기 이미지
            init_traj_img = env.get_visualized_img(img_choice='traj')
            init_obs_img = env.get_visualized_img(img_choice='obs')
            
            # 초기 빈 이미지 설치
            im_traj = axes[0].imshow(init_traj_img)
            im_obs = axes[1].imshow(init_obs_img)
            # im_pro_obs = axes[2].imshow(init_pro_obs_img)
            text_traj = axes[0].text(0.0, -0.15, "", transform=axes[0].transAxes, ha="left", fontsize=15, color='black')
            axes[0].set_title("Trajectory", fontsize=20)
            axes[1].set_title("Observation", fontsize=20)
            # axes[2].set_title("Processed Observation", fontsize=20)
            for ax in axes: ax.axis('off')
            plt.tight_layout()
            
            # Jupyter용 디스플레이 핸들을 생성 (이걸 통해 이미지만 쏙 바꿉니다)
            display_handle = display(fig, display_id=True)
            plt.close(fig) # 별도의 정적 출력이 생기지 않도록 닫기

        # 초기 상태(0 step)에서의 coverage 달성 여부 확인
        current_cov = env.coverage
        while target_idx < len(sorted_target_cov) and current_cov >= sorted_target_cov[target_idx]:
            target_cov_time_dict[sorted_target_cov[target_idx]] = (env.overlap_percent, env.cleaning_time)
            target_idx += 1
        
        # LSTM network에 통과시키면서 action을 얻으면서 경로 생성. Action 확률을 저장.
        while not done:
            
            # Network에 통과시켜 각 action이 선택될 확률을 출력
            loc_map_data = processed_obs["map"]
            loc_vec_data = processed_obs["loc_vec"]
            glob_vec_data = processed_obs["glob_vec"]
            action_mask = processed_obs["action_mask"]

            state_history.append({
                'h_t': h_t.clone(), 
                'c_t': c_t.clone(),
                'target_idx': target_idx,
                'target_cov_time_dict': target_cov_time_dict.copy()
            })

            with torch.no_grad():
                action_probs, h_t, c_t = self.policy_net(loc_map_data, loc_vec_data, glob_vec_data, h_t, c_t)
                
                # Training과 동일하게 log-probability 공간에서 mask를 적용한다.
                # +1e-8은 매우 작은 policy 확률이 log(0)이 되어 유효 action까지
                # 제거되는 것을 막고, mask=0인 action만 확실히 선택 불가하게 만든다.
                log_probs = torch.log(action_probs + 1e-8)
                masked_logits = torch.where(
                    action_mask == 1,
                    log_probs,
                    torch.tensor(-1e9, device=action_probs.device),
                )
            
            # Validation/test는 확률 sampling 없이, 유효 action 중 최대 확률 action을 결정적으로 선택.
            action = torch.argmax(masked_logits, dim=-1).item()

            if debug:
                if not done and debug_skip_count > 0:
                    debug_skip_count -= 1
                else:
                    
                    # 각 action이 선택될 확률 및 선택된 action 표시
                    before_mask_prob_np = action_probs.squeeze().cpu().numpy()
                    action_mask_np = action_mask.squeeze().cpu().numpy()
                    prob_np = torch.softmax(masked_logits, dim=-1).squeeze().cpu().numpy()

                    before_mask_probs_str = ", ".join([f"{prob:.3f}" for prob in before_mask_prob_np])
                    action_mask_str = ", ".join([f"{mask:.3f}" for mask in action_mask_np])
                    probs_str = ", ".join([f"{prob:.3f}" for prob in prob_np])

                    action_info = (f"[Selected action  ]: {action}\n"
                                   f"[Before mask probs]: [{before_mask_probs_str}]\n"
                                   f"[Action masks     ]: [{action_mask_str}]\n"
                                   f"[Action probs     ]: [{probs_str}]")
                    
                    # Map 시각화
                    traj_img = env.get_visualized_img(img_choice='traj')
                    obs_img = env.get_visualized_img(img_choice='obs')
                    
                    im_traj.set_data(traj_img)
                    im_obs.set_data(obs_img)
                    text_traj.set_text(action_info)
                    
                    # 화면 갱신 (도화지 위치는 그대로, 내용물만 부드럽게 변경)
                    display_handle.update(fig)

                    if not done: # Episode가 끝났으면 별도 입력 없이 바로 다음 map test 시작
                        user_val = input("Next step: [Enter] | Auto: [Number: If want to backstep, input negative number] | Exit: [q] >> ")
                        user_val = user_val.strip()

                        if user_val.lower() == 'q':
                            break
                        
                        try:
                            val = int(user_val)
                            if val < 0: # Back step 수행
                                backstep_num = abs(val)
                                last_obs = env.backstep(backstep_num)
                                processed_obs = self._pre_process_obs(last_obs)

                                pop_count = min(backstep_num, len(state_history))
                                for _ in range(pop_count):
                                    target_state = state_history.pop()
                                h_t = target_state['h_t']
                                c_t = target_state['c_t']
                                continue
                            elif val > 0: # Forward step 수행
                                debug_skip_count = int(user_val) - 1
                            else:
                                continue
                        except ValueError:
                            print("숫자 또는 'q'를 입력해주세요.")
            
            # Action 수행: 다음 obs 얻기
            next_obs, reward, terminated, truncated, info = env.step(action)
            processed_obs = self._pre_process_obs(next_obs, local_view_dim=LOCAL_VIEW_DIM)

            # Target Coverage 달성 여부 체크
            current_cov = env.coverage
            while target_idx < len(sorted_target_cov) and current_cov >= sorted_target_cov[target_idx]:
                target_cov_time_dict[sorted_target_cov[target_idx]] = (env.overlap_percent, env.cleaning_time)
                target_idx += 1
            
            # 각 환경별 종료 여부 판단
            done = terminated or truncated

        if debug:
            # Episode가 끝났을 때 map 시각화
            traj_img = env.get_visualized_img(img_choice='traj')
            obs_img = env.get_visualized_img(img_choice='obs')
            im_traj.set_data(traj_img)
            im_obs.set_data(obs_img)
            text_traj.set_text(action_info)
            display_handle.update(fig)
        
        coverage = env.coverage
        overlap_percent = env.overlap_percent
        cleaning_time = env.cleaning_time
        
        return coverage, overlap_percent, cleaning_time, target_cov_time_dict
    
    def train(self):
        
        assert self.args.mode == 'train'

        # Policy network를 train mode로 설정
        self.policy_net.train()        

        no_more_map_size = False
        try:
            curr_map_size = next(self.train_maps_sizes_iter)
        except StopIteration: # 마지막 map size인 경우, 이전에 설정한 map size를 그대로 사용
            no_more_map_size = True
            curr_map_size = self.train_map_sizes[-1]
        
        print(f"[Initial map size] {curr_map_size}.")
        
        # Episode 동안 map을 변경하지 않음. 고정된 map에서 경로 생성 수 policy update을 반복
        for episode in range(self.start_episode, self.train_cfg.max_episodes+1):

            if curr_map_size == '40x40':
                self.env_cfg.max_no_progress_steps = 30
                self.env_cfg.max_no_progress_steps_final = 10
                self.train_cfg.path_reward_thres = 0.85
            elif curr_map_size == '80x80':
                self.env_cfg.max_no_progress_steps = 45
                self.env_cfg.max_no_progress_steps_final = 20
                self.train_cfg.path_reward_thres = 0.75
            elif curr_map_size == 'Cropped':
                self.env_cfg.max_no_progress_steps = 60
                self.env_cfg.max_no_progress_steps_final = 30
                self.train_cfg.path_reward_thres = 0.75

            print(f"[Episode {episode}] Starting training...")
            eps_path_rwd_mean = 0 # Episode의 훈련 상태를 표시할 reward
            
            obs_list = []
            info_list = []
            
            # Map 설정
            for env_id in range(self.train_envs.num_envs):
                reset_info = None
                while not reset_info:
                    level = next(self.train_maps_level_iter[curr_map_size])
                    self.train_maps_level_iter_counts[curr_map_size] += 1
                    if self.train_cfg.use_train_maps:
                        map_file_path = next(self.train_maps[curr_map_size][level])
                        self.train_maps_iter_counts[curr_map_size][level] += 1
                        map_config = MapConfigSchema(file_path=map_file_path)
                    else:
                        H = None; W = None
                        if curr_map_size != "Cropped":
                            H, W = map(int, curr_map_size.split('x'))
                        map_config = MapConfigSchema(level=level, H=H, W=W)
                    reset_info = self.train_envs.call_at(method_name="reset", env_id=env_id, seed=None, map_config=map_config)
                obs, info = reset_info
                obs_list.append(obs)
                info_list.append(info)
            processed_obs = self._pre_process_obs(obs_list, local_view_dim=LOCAL_VIEW_DIM) # Map data를 resize, Observation data를 얻음
            
            steps = 0 # Episode 내에서의 step 수
            new_maps = True

            # 고정된 map에 대해서 경로 생성 최대 횟수를 넘으면 다른 map으로 바꿈. (while문을 빠져나옴)
            while steps < self.train_cfg.max_step_per_eps:
                
                # 1. Map 상태 reset: 
                # Map을 방금 새로 생성한 경우, reset이 이미 일어났으므로 reset을 생략.
                # 기존 map을 사용하는 경우, map 상태를 reset.
                if not new_maps:
                    obs_list = []
                    info_list = []
                    for env_id in range(self.train_envs.num_envs):
                        obs, info = self.train_envs.call_at(method_name="reset", env_id=env_id, seed=None, map_config=None)
                        obs_list.append(obs)
                        info_list.append(info)
                    processed_obs = self._pre_process_obs(obs_list, local_view_dim=LOCAL_VIEW_DIM)
                
                new_maps = False
                    
                # 2. Path 생성: LSTM network에 통과시켜서 각 action의 확률값을 얻고 map 상의 전체 경로 생성
                done = False
                
                # Hidden state와 cell state 초기화
                batch_size = len(obs_list)
                h_t = torch.zeros(batch_size, self.policy_net.lstm_hid_dim, device=self.device)
                c_t = torch.zeros(batch_size, self.policy_net.lstm_hid_dim, device=self.device)
                
                # log probability 저장
                saved_log_probs = []
                prob_masks = [] # 종료된 environment에서 계산된 log probability가 gradient update을 하지 않도록 하기 위한 mask
                saved_entropies = []
                curr_act_prob = np.ones(batch_size, dtype=bool) # 현재 아직 완료되지 않은 environment

                # LSTM network에 통과시키면서 action을 얻으면서 경로 생성. Action 확률을 저장.
                num_steps_in_map = 0
                while not done:

                    # Hidden state와 cell state를 detach하여 gradient가 LSTM step을 넘어서 전달되는 것을 방지.
                    if num_steps_in_map % self.train_cfg.detach_period == 0:
                        h_t = h_t.detach()
                        c_t = c_t.detach()
                    
                    # Network에 통과시켜 각 action이 선택될 확률을 출력
                    loc_map_data = processed_obs["map"]
                    loc_vec_data = processed_obs["loc_vec"]
                    glob_vec_data = processed_obs["glob_vec"]
                    action_mask = processed_obs["action_mask"]
                    action_probs, h_t, c_t = self.policy_net(loc_map_data, loc_vec_data, glob_vec_data, h_t, c_t)
                    
                    # Action이 나올 확률을 가지고 action 선택. Action 선택 확률을 저장. 장애물에 부딪히는 action은 제외
                    log_probs = torch.log(action_probs + 1e-8)
                    masked_logits = torch.where(action_mask == 1, log_probs, torch.tensor(-1e9, device=action_probs.device))

                    # Debugging용 코드
                    all_zero_per_env = (action_mask == 0).all(dim=-1)
                    if all_zero_per_env.any():
                        all_zero_env_indices = all_zero_per_env.nonzero(as_tuple=True)[0].tolist()
                        print(f"Action mask가 전부 0인 environment: {all_zero_env_indices}")
                        for i in all_zero_env_indices:
                            self.train_envs.call_at(method_name="show_visualized_img", env_id=i, img_choice='traj')

                    dist = torch.distributions.Categorical(logits=masked_logits)
                    actions_tensor = dist.sample() # Action을 확률을 기반으로 선택. Shape: (B,)
                    actions = actions_tensor.cpu().numpy() # step()을 위해 numpy 형식으로 변환
                    
                    # Entropy 계산: wandb logging용
                    entropy = dist.entropy().detach()
                    saved_entropies.append(entropy)
                    
                    # Action 수행: 다음 obs 얻기
                    next_obs_list, rewards, terminateds, truncateds, info = self.train_envs.step(actions)
                    processed_obs = self._pre_process_obs(next_obs_list, local_view_dim=LOCAL_VIEW_DIM)
                    
                    # 확률 미분 계산을 위해 log probability를 저장
                    selected_log_probs = dist.log_prob(actions_tensor)
                    saved_log_probs.append(selected_log_probs)
                    prob_masks.append(torch.as_tensor(curr_act_prob, device=self.device, dtype=torch.float32))
                    
                    # 각 환경별 종료 여부 판단
                    dones = terminateds | truncateds # Shape: (B,)
                    done = dones.all()
                    curr_act_prob = curr_act_prob & ~dones # 이미 종료된 environment는 0으로 표시 -> Gradient 계산 시 probability가 영향을 주지 않도록 함.

                    num_steps_in_map += 1

                ##############################################################
                # [PATH REWARD FUNCTION]
                ##############################################################
                # 각 environment에서 얻은 path의 coverage, overlap rate, cleaning time 계산
                coverages = np.array(self.train_envs.get_attr("coverage")) # Shape: (B,)
                overlap_rates = np.array(self.train_envs.get_attr("overlap_percent")) / 100 # Shape: (B,)
                cleaning_times = np.array(self.train_envs.get_attr("cleaning_time")) # Shape: (B,)
                
                # cleaning time의 기준 설정: Total step을 직진으로 이동했을 때 걸린 시간을 기준으로 삼음.
                total_steps = np.array(self.train_envs.get_attr("steps")) # Shape: (B,)
                straight_cleaning_times = total_steps * GRID_RESOLUTION_M / LINEAR_VELOCITY / 60.0
                cleaning_times_rel = cleaning_times / (straight_cleaning_times + 1e-8) # 경로 길이에 대한 cleaning time
                
                # Tensor 변환
                cov_tensor = torch.tensor(coverages, device=self.device, dtype=torch.float32).detach()
                overlap_tensor = torch.tensor(overlap_rates, device=self.device, dtype=torch.float32).detach()
                time_tensor = torch.tensor(cleaning_times_rel, device=self.device, dtype=torch.float32).detach()
                
                path_reward = self._calculate_path_reward(
                    cov_tensor, overlap_tensor, time_tensor, curr_map_size
                )
                
                advantage = (path_reward - path_reward.mean()).detach() # Shape: (B,)
                ##############################################################
                
                # loss 계산 및 parameter update
                log_probs_tensor = torch.stack(saved_log_probs) # Shape: (T, B) (T: time steps)
                prob_masks_tensor = torch.stack(prob_masks) # Shape: (T, B)
                steps_tensor = prob_masks_tensor.sum(dim=0) # Shape: (B,)
                # log probability를 합하여 전체 경로의 probability 계산 
                # (완료된 environment에서 얻은 probability는 계수 0을 곱하여 gradient 전달이 안 되도록 함.)
                sum_log_probs = (log_probs_tensor * prob_masks_tensor).sum(dim=0) # Shape: (B,)
                mean_log_probs = sum_log_probs / (steps_tensor + 1e-8) # Shape: (B,)
                policy_loss = -(mean_log_probs * advantage) # 음수를 붙여서 path_reward가 증가하는 방향으로 update 되도록 함.
                
                loss = policy_loss.mean() # Batch 평균 loss
                self.optimizer.zero_grad() # Gradient 초기화
                loss.backward() # Gradient 계산
                grad_norm_before = torch.nn.utils.clip_grad_norm_(self.policy_net.parameters(), max_norm=1.0) # Gradient clipping
                self.optimizer.step() # Parameter update
                
                # Entropy 통계
                entropies_tensor = torch.stack(saved_entropies) # Shape: (T, B)
                mean_entropy = (entropies_tensor * prob_masks_tensor).sum() / (prob_masks_tensor.sum() + 1e-8)
                
                # Path reward의 평균
                path_reward_mean = path_reward.mean().item()
                eps_path_rwd_mean = path_reward_mean # Episode path reward update (가장 최근 model의 path reward를 출력할 것임.)
                
                log_dict = {
                    "train/episode": episode,
                    "train/step": steps,
                    "train/loss": loss.item(),
                    "train/grad_norm_raw": grad_norm_before.item() if isinstance(grad_norm_before, torch.Tensor) else grad_norm_before,
                    "train/entropy_mean": mean_entropy.item(), # 💡 수집된 평균 엔트로피
                    "train/map_steps_mean": total_steps.mean(),
                    "train/learning_rate": self.optimizer.param_groups[0]['lr'],
                    "train/total_steps": self.total_steps,
                    
                    # Reward 지표 (Mean, Max, Min)
                    "train/path_reward_mean": path_reward_mean,
                    "train/path_reward_max": path_reward.max().item(),
                    "train/path_reward_min": path_reward.min().item(),
                    
                    # Coverage 지표 (Mean, Max, Min)
                    "train/coverage_mean": cov_tensor.mean().item(),
                    "train/coverage_max": cov_tensor.max().item(),
                    "train/coverage_min": cov_tensor.min().item(),
                    
                    # Overlap 지표 (Mean, Max, Min)
                    "train/overlap_mean": overlap_tensor.mean().item(),
                    "train/overlap_max": overlap_tensor.max().item(),
                    "train/overlap_min": overlap_tensor.min().item(),
                    
                    # Time 지표
                    "train/time_mean": cleaning_times.mean(),
                    "train/time_ratio_mean": time_tensor.mean().item(),
                }
                
                # wandb가 초기화되어 있다면 기록
                if self.args.use_wandb: # 혹은 관련 조건문
                    self.wandb_run.log(log_dict, step=self.total_steps)
                
                # Steps: Parameter update 횟수
                steps += 1
                self.total_steps += 1
                
            # 주어진 map size에서 training한 episode 횟수 증가
            self.eps_num_per_map_size += 1
            
            # 좋은 경로 생성에 성공했거나 episode 횟수가 한계치를 넘어서면 map size를 바꿈
            change_map_size = False
            if not no_more_map_size and self.eps_num_per_map_size > self.train_cfg.max_eps_per_map_size: # 변경할 map_size가 더 없는 경우에는 change_map_size를 무조건 False로 설정되도록 함.
                change_map_size = True
                print(f"[Map Scale Up] Reached maximum episodes ({self.eps_num_per_map_size}) for map size {curr_map_size}.")
            
            # 한 map에 대해서 훈련이 끝나면 checkpoint 저장 및 validation 수행

            # Episode 훈련 결과 출력
            print(f"\tEpisode reward: {eps_path_rwd_mean}")
                
            # Validation 수행
            if episode % self.train_cfg.valid_freq == 0:
                print("Validation...")
                start_mode = "corner" if curr_map_size == "Cropped" else "edge"
                valid_path_reward_mean = self._validation(episode, curr_map_size, start_mode)
                if valid_path_reward_mean >= self.train_cfg.path_reward_thres:
                    change_map_size = True
                    print(f"[Map Scale Up] Achieved path reward {valid_path_reward_mean:.3f} at map size {curr_map_size} on validation. "
                          f"Exceeded the threshold {self.train_cfg.path_reward_thres}.")

                
            if change_map_size:
                try:
                    curr_map_size = next(self.train_maps_sizes_iter)
                    self.train_maps_sizes_iter_count += 1
                except StopIteration: # 마지막 map size인 경우, 이전에 설정한 map size를 그대로 사용
                    print("[Map Scale Up] No more map sizes available. Maintaining the current size.")
                    curr_map_size = self.train_map_sizes[-1]
                    no_more_map_size = True
                else:
                    print(f"[Map Scale Up] Successfully changed map size to {curr_map_size}.")
                    self.eps_num_per_map_size = 0
            
            # Map 기록 저장
            if episode % 5 == 0:
                map_img = self.train_envs.call_at(method_name="get_visualized_img", env_id=0, img_choice="traj")
                if self.wandb_run:
                    self.wandb_run.log({"Visualization/Robot_path": wandb.Image(map_img)}, step=self.total_steps)

            self.scheduler.step()

            # Checkpoint 저장
            if episode % self.train_cfg.ckp_freq == 0:
                checkpoint_info = {"episode": episode}
                self._save_model(mode='checkpoint', info=checkpoint_info)

    def test(self, target_cov_list: list[float] = [0.85, 0.90, 0.95]):

        self.policy_net.eval() # eval mode로 전환

        print_model_info(self.policy_net)

        total_coverage = []; total_overlap_percent = []; total_cleaning_time = []
        computation_time = []
        map_result_rows: list[dict] = []
        result_dir = self._get_test_result_dir()

        sorted_target_cov = sorted(target_cov_list)

        # Level별로 coverage 결과를 저장하기 위한 dictionary
        # coverage_results_per_level = {
        #     level: [coverage_list],
        #     ...  
        # }
        coverage_results_per_level: dict[int, list[float]] = {}

        # Target coverage에 따른 결과를 저장하기 위한 dictionary
        # target_cov_results = {
        #     target_cov: {
        #         level: {
        #             'total_count': total_count,
        #             'success_count': success_count,
        #             'overlaps': [overlap_list],
        #             'cleaning_times': [cleaning_time_list]
        #         }
        #         ...
        #     }
        #     ...
        # }
        target_cov_results: dict[float, dict[int, dict[str, int | list[float]]]] = {
            cov: {} for cov in sorted_target_cov
        }
        
        # Level별 Best/Worst 경로 기록용 딕셔너리
        visualized_maps: dict[int, dict[str, list[tuple[tuple[float, float, float], np.ndarray, str]]]] = {}

        map_num_per_level = self.args.test_map_num_per_level

        print(f"Use {map_num_per_level} maps per level for test.")
        
        # 각 map에 대해서 test
        for level in self.test_maps.keys():

            level_records = []  # Level의 모든 결과를 저장하기 위한 list
            
            # Dictionary 구조 초기화
            if level not in coverage_results_per_level:
                coverage_results_per_level[level] = []
                
            for cov in sorted_target_cov:
                if level not in target_cov_results[cov]:
                    target_cov_results[cov][level] = {
                        'total_count': 0,
                        'success_count': 0,
                        'overlaps': [],
                        'cleaning_times': []
                    }

            map_indices = range(min(map_num_per_level, len(self.test_maps[level])))
            for map_idx in tqdm(map_indices, desc=f"Level {level} Maps"):

                map_path = self.test_maps[level][map_idx]
                map_config = MapConfigSchema(file_path=map_path)
                map_name = os.path.basename(map_path)

                reset_info = self.test_env.reset(seed=None, start_mode='corner', map_config=map_config)
                if not reset_info:
                    print(f"Can't reset map file: {os.path.basename(self.test_maps[level][map_idx])}")
                    continue
                obs, _ = reset_info
                
                # 여러 starting point에서 model 성능을 test
                for start_idx in range(self.args.test_start_point_num):
                
                    if start_idx != 0:
                        obs, _ = self.test_env.reset(start_mode='corner')  # 시작 지점 초기화
                    
                    start_time = time.time()
                    cur_coverage, cur_overlap_percent, cur_cleaning_time, target_cov_time_dict = self._test_one_map(
                        self.test_env, obs, target_cov_list=sorted_target_cov, debug=self.args.debug
                    ) 
                    end_time = time.time()
                    comp_t = end_time - start_time
                    computation_time.append(comp_t)
                    
                    # Reachable grid의 수가 전체 grid 수의 절반을 넘는 경우에만 결과 저장
                    if self.test_env.map_layers.coverable.sum() >= self.test_env.H * self.test_env.W * 0.5:
                        
                        # 1) Level별 Final Coverage 저장
                        coverage_results_per_level[level].append(cur_coverage)
                        
                        # 2) Target Coverage별 지표 저장
                        for cov in sorted_target_cov:
                            t_overlap, t_time = target_cov_time_dict.get(cov, (-1.0, -1.0))
                            
                            target_lvl_data = target_cov_results[cov][level]
                            target_lvl_data['total_count'] += 1
                            
                            # 성공한 경우 (-1이 아닌 경우)
                            if t_time != -1.0 and t_time > 0:
                                target_lvl_data['success_count'] += 1
                                target_lvl_data['overlaps'].append(t_overlap)
                                target_lvl_data['cleaning_times'].append(t_time)
                        
                        # 시각화용 데이터 레코드 생성
                        cand_metrics = (cur_coverage, cur_overlap_percent, cur_cleaning_time)
                        cand_img = self.test_env.get_visualized_img(img_choice='traj')
                        cand_info = f"Map name: {os.path.basename(map_path)} | Final Cov: {cur_coverage*100:.1f}%, Overlap: {cur_overlap_percent:.1f}%, Time: {cur_cleaning_time:.1f}m"

                        level_records.append((cand_metrics, cand_img, cand_info))

                        map_result_rows.append({
                            'coverage': cur_coverage,
                            'time': cur_cleaning_time,
                            'overlap': cur_overlap_percent,
                            'target_metrics': target_cov_time_dict,
                            'map_name': map_name,
                        })
                        self._save_test_path_image(result_dir, map_name, cand_img)

            # Best / Median / Worst 경로 선별
            if len(level_records) > 0:
                import functools
                
                def compare_paths(item1, item2):
                    metrics1, metrics2 = item1[0], item2[0]
                    if is_better_path(metrics1, metrics2):
                        return -1
                    elif is_better_path(metrics2, metrics1):
                        return 1
                    return 0

                sorted_records = sorted(level_records, key=functools.cmp_to_key(compare_paths))
                total_cnt = len(sorted_records)
                num_sample = self.args.vis_test_map_num

                bests = sorted_records[:min(num_sample, total_cnt)]
                worsts = sorted_records[-min(num_sample, total_cnt):][::-1]

                mid_idx = total_cnt // 2
                half_k = num_sample // 2
                start_m = max(0, mid_idx - half_k)
                end_m = min(total_cnt, start_m + num_sample)
                medians = sorted_records[start_m:end_m]

                visualized_maps[level] = {
                    'best': bests,
                    'median': medians,
                    'worst': worsts
                }

        # =========================================================================
        # 1. Level별 Final Coverage 출력
        # =========================================================================
        print("\n" + "="*65)
        print("[Final Coverage Summary by Level]")
        print("="*65)
        
        for level in sorted(coverage_results_per_level.keys()):
            covs = coverage_results_per_level[level]
            if covs:
                cov_m, cov_s = np.mean(covs), np.std(covs)
                print(f" ▶ Level {level} (Total tests: {len(covs)})")
                print(f"    - Final Coverage(%): {cov_m*100:.2f} ± {cov_s*100:.2f}%")
            else:
                print(f" ▶ Level {level} -> No Valid Data")

        # =========================================================================
        # 2. Target Coverage별 & Level별 세부 및 Overall 성능 출력
        # =========================================================================
        print("\n" + "="*65)
        print("[2. Detailed Results by Target Coverage]")
        print("="*65)
        
        for cov in sorted_target_cov:
            print(f"\n Target Coverage: {cov*100:.1f}%")
            print("-" * 60)
            
            total_attempts = 0
            total_successes = 0
            all_overlaps = []
            all_times = []
            
            for level in sorted(target_cov_results[cov].keys()):
                lvl_data = target_cov_results[cov][level]
                t_cnt = lvl_data['total_count']
                s_cnt = lvl_data['success_count']
                
                total_attempts += t_cnt
                total_successes += s_cnt
                all_overlaps.extend(lvl_data['overlaps'])
                all_times.extend(lvl_data['cleaning_times'])
                
                if t_cnt == 0:
                    print(f"   Level {level}: No Data")
                    continue
                
                sr = (s_cnt / t_cnt) * 100.0
                
                if s_cnt > 0:
                    ov_m, ov_s = np.mean(lvl_data['overlaps']), np.std(lvl_data['overlaps'])
                    tm_m, tm_s = np.mean(lvl_data['cleaning_times']), np.std(lvl_data['cleaning_times'])
                    print(f"   Level {level} | Success Rate: {sr:5.1f}% ({s_cnt}/{t_cnt}) "
                        f"| Overlap: {ov_m:5.2f} ± {ov_s:4.2f}% "
                        f"| Time: {tm_m:5.2f} ± {tm_s:4.2f}m")
                else:
                    print(f"   Level {level} | Success Rate:   0.0% ({s_cnt}/{t_cnt}) "
                        f"| Overlap: N/A | Time: N/A")
            
            # Overall 출력
            print("-" * 60)
            if total_attempts > 0:
                overall_sr = (total_successes / total_attempts) * 100.0
                if total_successes > 0:
                    ov_m, ov_s = np.mean(all_overlaps), np.std(all_overlaps)
                    tm_m, tm_s = np.mean(all_times), np.std(all_times)
                    print(f"   [Overall for Target {cov*100:.1f}%]")
                    print(f"   - Success Rate : {overall_sr:.2f}% ({total_successes}/{total_attempts})")
                    print(f"   - Overlap Rate : {ov_m:.2f} ± {ov_s:.2f}%")
                    print(f"   - Cleaning Time: {tm_m:.2f} ± {tm_s:.2f} min")
                else:
                    print(f"   [Overall for Target {cov*100:.1f}%]")
                    print(f"   - Success Rate : 0.00% ({total_successes}/{total_attempts})")
                    print(f"   - Overlap Rate : N/A | Cleaning Time: N/A")

        # 평균 연산 시간 출력
        avg_comp_time = np.mean(computation_time) if computation_time else 0.0
        median_comp_time = np.median(computation_time) if computation_time else 0.0
        print("\n" + "="*65)
        print(f"Average computation time per map: {avg_comp_time:.2f} s")
        print(f"Median computation time per map: {median_comp_time:.2f} s")
        print("="*65)

        result_file_path = self._save_test_result_workbook(
            result_dir,
            map_result_rows,
            sorted_target_cov,
            target_cov_results,
        )
        print(f"Test results saved to: {result_file_path}")

        # =========================================================================
        # Level별 경로 시각화 출력
        # =========================================================================
        for level in sorted(visualized_maps.keys()):
            vis_data = visualized_maps[level]
            categories = [
                ('best', '[Best Paths]'),
                ('median', '[Median Paths]'),
                ('worst', '[Worst Paths]')
            ]
            
            print(f"\n" + "="*50)
            print(f" Visualization for Map Condition Level {level}")
            print("="*50)
            
            for cat_key, cat_label in categories:
                records = vis_data.get(cat_key, [])
                if records:
                    print(f"\n  {cat_label} (Top {len(records)})")
                    for idx, (metrics, img, info_str) in enumerate(records, 1):
                        print(f"    #{idx} - {info_str}")
                        display_image(img)

    def _get_test_result_dir(self) -> str:
        """Return the model-specific directory used for persisted test results."""
        model_file_name = os.path.basename(self.args.model_name)
        model_name, _ = os.path.splitext(model_file_name)
        result_dir = os.path.join(RESULT_SAVE_DIR, model_name)
        os.makedirs(result_dir, exist_ok=True)
        return result_dir

    @staticmethod
    def _save_test_path_image(result_dir: str, map_name: str, rgb_image: np.ndarray) -> None:
        """Save the environment's RGB trajectory image as ``<map>_path.png``."""
        map_stem, _ = os.path.splitext(map_name)
        image_path = os.path.join(result_dir, f"{map_stem}_path.png")
        bgr_image = cv2.cvtColor(rgb_image, cv2.COLOR_RGB2BGR)
        if not cv2.imwrite(image_path, bgr_image):
            raise IOError(f"Failed to save test path image: {image_path}")

    def _save_test_result_workbook(
        self,
        result_dir: str,
        map_result_rows: list[dict],
        target_cov_list: list[float],
        target_cov_results: dict[float, dict[int, dict[str, int | list[float]]]],
    ) -> str:
        """Write per-map metrics and level/target summaries to one Excel file."""

        workbook = Workbook()
        map_sheet = workbook.active
        map_sheet.title = "map_results"

        map_headers = ["No.", "Coverage(%)", "Time(min)", "Overlap(%)"]
        for target_cov in target_cov_list:
            target_name = f"Coverage {target_cov * 100:g}%"
            map_headers.extend([
                f"{target_name} reached",
                f"{target_name} time",
                f"{target_name} overlap",
            ])
        map_headers.append("Map name")
        map_sheet.append(map_headers)

        for index, row_data in enumerate(map_result_rows, start=1):
            row = [
                index,
                row_data['coverage'] * 100,
                row_data['time'],
                row_data['overlap'],
            ]
            target_metrics = row_data['target_metrics']
            for target_cov in target_cov_list:
                target_overlap, target_time = target_metrics.get(target_cov, (-1.0, -1.0))
                reached = target_time >= 0.0
                row.extend([
                    "TRUE" if reached else "FALSE",
                    target_time if reached else None,
                    target_overlap if reached else None,
                ])
            row.append(row_data['map_name'])
            map_sheet.append(row)

        summary_sheet = workbook.create_sheet("summary")
        summary_sheet.append([
            "Level",
            "Target coverage",
            "Success rate",
            "Time mean (median, std)",
            "Overlap mean (median, std)",
        ])
        levels = sorted({
            level
            for target_data in target_cov_results.values()
            for level in target_data
        })
        for level in levels:
            for target_cov in target_cov_list:
                level_data = target_cov_results[target_cov][level]
                total_count = int(level_data['total_count'])
                success_count = int(level_data['success_count'])
                success_rate = success_count / total_count if total_count else 0.0
                times = level_data['cleaning_times']
                overlaps = level_data['overlaps']
                if success_count:
                    time_text = self._format_summary_metric(times, "min")
                    overlap_text = self._format_summary_metric(overlaps, "%")
                else:
                    time_text = "N/A"
                    overlap_text = "N/A"
                summary_sheet.append([
                    level,
                    target_cov,
                    success_rate,
                    time_text,
                    overlap_text,
                ])

        for sheet in (map_sheet, summary_sheet):
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for column_cells in sheet.columns:
                column_letter = column_cells[0].column_letter
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

        for row in map_sheet.iter_rows(min_row=2, min_col=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"
        for row in summary_sheet.iter_rows(min_row=2, min_col=2, max_col=3):
            row[1].number_format = "0.00"
            row[0].number_format = "0"

        workbook_path = os.path.join(result_dir, "test_results.xlsx")
        workbook.save(workbook_path)
        return workbook_path

    @staticmethod
    def _format_summary_metric(values: list[float], unit: str) -> str:
        values_array = np.asarray(values, dtype=float)
        return (
            f"{np.mean(values_array):.2f} "
            f"({np.median(values_array):.2f}, {np.std(values_array):.2f}) {unit}"
        )

    def see_weight(self):

        import torch.nn as nn

        model = self.policy_net
        cmps_weight = model.cmps_net.weight.detach().cpu().numpy()

        map_dim = model.map_feat_dim
        vec_dim = model.vec_feat_dim
        lstm_dim = model.lstm_hid_dim

        # =========================================================================
        # 1. cmps_net 가중치 지분율 시각화 (Total Weight Share - Sum 기준)
        # =========================================================================
        w_map = cmps_weight[:, :map_dim]
        w_vec = cmps_weight[:, map_dim : map_dim + vec_dim]
        w_lstm = cmps_weight[:, map_dim + vec_dim :]

        # 각 파트별 가중치 절대값의 '총합' (진짜 지분율)
        imp_map = np.abs(w_map).sum()
        imp_vec = np.abs(w_vec).sum()
        imp_lstm = np.abs(w_lstm).sum()

        total = imp_map + imp_vec + imp_lstm
        ratios = [(imp_map / total) * 100, (imp_vec / total) * 100, (imp_lstm / total) * 100]
        labels = [f'Map Feat\n({map_dim}d)', f'Vector Feat\n({vec_dim}d)', f'LSTM State\n({lstm_dim}d)']

        fig1, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5), gridspec_kw={'width_ratios': [2, 1]})

        # [왼쪽] cmps_net Weight Matrix 전체 히트맵
        im = ax1.imshow(cmps_weight, cmap='seismic', aspect='auto', vmin=-np.max(np.abs(cmps_weight)), vmax=np.max(np.abs(cmps_weight)))
        ax1.set_title("cmps_net Weight Matrix Heatmap", fontsize=12, fontweight='bold')
        ax1.set_xlabel("Input Feature Dimensions (Map | Vector | LSTM)")
        ax1.set_ylabel("Output Dimension (action_enc_dim)")

        # 그룹별 경계선(구분선) 그리기
        ax1.axvline(x=map_dim - 0.5, color='black', linestyle='--', linewidth=1.5)
        ax1.axvline(x=map_dim + vec_dim - 0.5, color='black', linestyle='--', linewidth=1.5)

        cbar = fig1.colorbar(im, ax=ax1)
        cbar.set_label('Weight Value')

        # [오른쪽] 그룹별 총 지분율 Bar Chart
        colors = ['#4C72B0', '#55A868', '#C44E52']
        bars = ax2.bar(labels, ratios, color=colors, alpha=0.85, width=0.5)
        ax2.set_title("Feature Group Importance Ratio (%)", fontsize=12, fontweight='bold')
        ax2.set_ylabel("Total Weight Share Ratio (%)")
        ax2.set_ylim(0, max(ratios) * 1.2)

        for bar, pct in zip(bars, ratios):
            yval = bar.get_height()
            ax2.text(bar.get_x() + bar.get_width()/2.0, yval + 1.0, f'{pct:.1f}%', ha='center', va='bottom', fontweight='bold')

        plt.tight_layout()
        plt.show()

        
        # =========================================================================
        # 2. 최초 입력 3개 Map/Channel별 가중치(Weight) 비중 분석
        # =========================================================================
        # map_enc의 첫 번째 Conv2d 레이어 찾기
        first_conv = None
        for layer in model.map_enc:
            if isinstance(layer, nn.Conv2d):
                first_conv = layer
                break

        # w_first Shape: (32, 3, 3, 3) -> (out_channels, in_channels, K_h, K_w)
        w_first = first_conv.weight.detach().cpu().numpy()
        in_channels = w_first.shape[1]  # 보통 3 (또는 3 * stack_steps)

        # 입력 채널별 가중치 절대값 합(Sum) 및 평균(Mean) 계산
        ch_sums = [np.abs(w_first[:, c, :, :]).sum() for c in range(in_channels)]
        ch_means = [np.abs(w_first[:, c, :, :]).mean() for c in range(in_channels)]

        total_ch_sum = sum(ch_sums)
        ch_ratios = [(s / total_ch_sum) * 100 for s in ch_sums]

        # 입력 채널 라벨 설정 (필요시 이름 수정)
        ch_labels = [f'Input Ch {i+1}' for i in range(in_channels)]

        print("=" * 60)
        print("📊 [First Conv2d Layer - Input Channel Weight Analysis]")
        print("=" * 60)
        for i in range(in_channels):
            print(f"  - {ch_labels[i]} : Total Share = {ch_ratios[i]:5.2f}% | Mean |W| = {ch_means[i]:.6f}")
        print("=" * 60)

        # 시각화 (입력 채널별 가중치 비율 Bar Chart)
        fig3, (ax6, ax7) = plt.subplots(1, 2, figsize=(11, 4))
        ch_colors = ['#4C72B0', '#55A868', '#C44E52']

        # [좌] 입력 채널별 총 가중치 지분율 (%)
        bars6 = ax6.bar(ch_labels, ch_ratios, color=ch_colors[:in_channels], alpha=0.85, width=0.4)
        ax6.set_title("Input Map Channels Weight Share Ratio (%)", fontsize=11, fontweight='bold')
        ax6.set_ylabel("Weight Share (%)")
        ax6.set_ylim(0, max(ch_ratios) * 1.25)
        for bar, pct in zip(bars6, ch_ratios):
            ax6.text(bar.get_x() + bar.get_width()/2.0, bar.get_height() + 1.0, f'{pct:.1f}%', ha='center', va='bottom', fontweight='bold')

        # [우] 입력 채널별 커널 파라미터 평균 크기
        bars7 = ax7.bar(ch_labels, ch_means, color=ch_colors[:in_channels], alpha=0.85, width=0.4)
        ax7.set_title("Input Map Channels Mean |Weight|", fontsize=11, fontweight='bold')
        ax7.set_ylabel("Mean Magnitude")
        ax7.set_ylim(0, max(ch_means) * 1.25)
        for bar, val in zip(bars7, ch_means):
            ax7.text(bar.get_x() + bar.get_width()/2.0, bar.get_height() + 0.0005, f'{val:.4f}', ha='center', va='bottom', fontweight='bold')

        plt.tight_layout()
        plt.show()

    def test_one_map_for_debug(self, map_rel_path: str):

        map_full_path = os.path.join(self.args.map_save_dir, map_rel_path)
        map_config = MapConfigSchema(file_path=map_full_path)
        map_name = os.path.basename(map_full_path)

        reset_info = self.test_env.reset(seed=None, start_mode='edge', map_config=map_config)
        if not reset_info:
            print(f"Can't reset map file: {map_name}")
            return
        obs, _ = reset_info
        _, _, _, _ = self._test_one_map(self.test_env, obs, debug=True) 
            
